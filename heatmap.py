import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import numpy as np
import io
import json
from datetime import datetime
from pathlib import Path
import altair as alt
import os
import uuid
import base64

# ----------- Configuration -----------
APP_VERSION = "1.1.0"
CONFIG_FILE = "config.json"
DATA_FOLDER = "data"
DATA_FILE = os.path.join(DATA_FOLDER, "data.csv")
BACKUP_FOLDER = os.path.join(DATA_FOLDER, "backups")
COLUMNS = [
    "ID", "Opportunity", "Related to", "Area", "Type", "Topic",
    "Impact", "Complexity", "Score", "Status", "Created", "Modified"
]
STATUS_OPTIONS = ["Idea", "To explore", "Validated", "In development", "Deployed"]
TYPE_OPTIONS = ["Enabler", "Lever"]

# ----------- Setup -----------
def initialize_folders():
    """Create necessary folders if they don't exist"""
    for folder in [DATA_FOLDER, BACKUP_FOLDER]:
        os.makedirs(folder, exist_ok=True)

def load_config():
    """Load configuration from JSON file"""
    config_path = Path(CONFIG_FILE)
    default_config = {
        "language": "fr",
        "theme": "light",
        "backup_frequency": 5,  # Number of changes before auto backup
        "custom_areas": [],
        "custom_topics": []
    }
    
    if config_path.exists():
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            st.error(f"Error loading config: {e}")
            return default_config
    else:
        # Create default config file
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(default_config, f, indent=4)
        return default_config

def save_config(config):
    """Save configuration to JSON file"""
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=4)

# ----------- Translation System -----------
TRANSLATIONS = {
    "en": {
        "app_title": "🎯 Priority Heat Map for AI Opportunities",
        "navigation": "Navigation",
        "dashboard": "Dashboard",
        "management": "Management",
        "visualization": "Visualization",
        "export": "Export",
        "settings": "Settings",
        # Add more translations as needed
        "add_new": "Add New Opportunity",
        "edit": "Edit Opportunity",
        "delete": "Delete Opportunity",
        "confirm_delete": "Confirm Delete"
    },
    "fr": {
        "app_title": "🎯 Carte de chaleur de priorité pour les opportunités IA",
        "navigation": "Navigation",
        "dashboard": "Tableau de bord",
        "management": "Gestion",
        "visualization": "Visualisation",
        "export": "Export",
        "settings": "Paramètres",
        # Add more translations as needed
        "add_new": "Ajouter une Nouvelle Opportunité",
        "edit": "Modifier une Opportunité",
        "delete": "Supprimer une Opportunité",
        "confirm_delete": "Confirmer la Suppression"
    }
}

def get_text(key, lang):
    """Get translated text"""
    try:
        return TRANSLATIONS[lang][key]
    except KeyError:
        # Fallback to English if translation not found
        try:
            return TRANSLATIONS["en"][key]
        except KeyError:
            return key  # Return the key itself if no translation exists

# ----------- Data Management Functions -----------
@st.cache_data(ttl=300)
def compute_score(impact, complexity):
    """Calculate priority score"""
    return round((impact + (10 - complexity)) / 2, 1)

def create_backup(df):
    """Create a backup of the current data"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = os.path.join(BACKUP_FOLDER, f"data_backup_{timestamp}.csv")
    df.to_csv(backup_file, index=False)
    
    # Keep only the 10 most recent backups
    backups = sorted(Path(BACKUP_FOLDER).glob("*.csv"), key=os.path.getmtime, reverse=True)
    for old_backup in backups[10:]:
        try:
            old_backup.unlink()
        except:
            pass

def load_data():
    """Load data from CSV file with error handling"""
    try:
        if Path(DATA_FILE).exists():
            df = pd.read_csv(DATA_FILE)
            # Ensure all required columns exist
            for col in COLUMNS:
                if col not in df.columns:
                    if col in ["Created", "Modified"]:
                        df[col] = datetime.now().strftime("%Y-%m-%d %H:%M")
                    else:
                        df[col] = None
            
            # Convert string columns that should be numeric
            for col in ["Impact", "Complexity", "Score"]:
                df[col] = pd.to_numeric(df[col], errors='coerce').round(1)
                
            # Fill any NaN values
            df = df.fillna({
                "ID": 0,
                "Opportunity": "",
                "Related to": "",
                "Area": "",
                "Type": TYPE_OPTIONS[0],
                "Topic": "",
                "Impact": 5.0,
                "Complexity": 5.0,
                "Score": 5.0,
                "Status": STATUS_OPTIONS[0]
            })
            
            return df
        else:
            return pd.DataFrame(columns=COLUMNS)
    except Exception as e:
        st.error(f"Error loading data: {e}")
        return pd.DataFrame(columns=COLUMNS)

def save_data(df):
    """Save data to CSV file with error handling"""
    try:
        # Create folder if not exists
        os.makedirs(os.path.dirname(DATA_FILE), exist_ok=True)
        
        # Save data
        df.to_csv(DATA_FILE, index=False)
        
        # Initialize save_counter if not exists
        if "save_counter" not in st.session_state:
            st.session_state.save_counter = 0
        
        st.session_state.save_counter += 1
        
        # Create backup if needed
        config = st.session_state.get("config", {"backup_frequency": 5})
        backup_frequency = config.get("backup_frequency", 5)
        
        if st.session_state.save_counter >= backup_frequency:
            create_backup(df)
            st.session_state.save_counter = 0
            
        return True
    except Exception as e:
        st.error(f"Error saving data: {e}")
        return False

def refresh_data(df):
    """Sort data by score and update IDs"""
    df = df.sort_values(by="Score", ascending=False).reset_index(drop=True)
    df["ID"] = range(1, len(df) + 1)
    df["Modified"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    return df

def generate_unique_id():
    """Generate a unique identifier for opportunities"""
    return str(uuid.uuid4())[:8]

# ----------- UI Helpers -----------
def create_color_scale():
    """Create a color scale function based on scores"""
    def get_color(score):
        if score >= 8:
            return "#c6f5d3"  # Green
        elif score >= 6:
            return "#fff3b3"  # Yellow
        else:
            return "#f8d3d3"  # Red
    return get_color

def add_logo():
    """Add a small logo to the sidebar (if available)"""
    logo_path = "logo.png"
    if Path(logo_path).exists():
        st.sidebar.image(logo_path, width=100)

def display_score_indicator(score, size=100):
    """Display a circular indicator for the score with improved styling"""
    fig, ax = plt.subplots(figsize=(size/100, size/100))
    
    # Set figure facecolor to transparent
    fig.patch.set_alpha(0.0)
    
    # Set the background circle
    background = plt.Circle((0.5, 0.5), 0.45, color='none')
    
    # Set the score circle color based on value
    if score >= 8:
        color = '#c6f5d3'  # Green
    elif score >= 6:
        color = '#fff3b3'  # Yellow
    else:
        color = '#f8d3d3'  # Red
    
    score_circle = plt.Circle((0.5, 0.5), 0.45, color=color)
    
    ax.add_patch(background)
    ax.add_patch(score_circle)
    
    # Add the score text
    plt.text(0.5, 0.5, f"{score}", 
             horizontalalignment='center',
             verticalalignment='center',
             fontsize=size/6, fontweight='bold', color='black')
    
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    plt.axis('off')
    plt.tight_layout()
    
    return fig

def get_download_link(data, filename, text):
    """Generate a download link for data"""
    b64 = base64.b64encode(data).decode()
    href = f'<a href="data:file/csv;base64,{b64}" download="{filename}">{text}</a>'
    return href

# ----------- Page Functions -----------
def dashboard_page(data, lang="en"):
    """Render the dashboard page"""
    st.header("📊 " + get_text("dashboard", lang))
    
    # KPIs
    if not data.empty:
        # Top stats in cards with colored backgrounds
        col1, col2, col3, col4 = st.columns(4)
        
        # Calculate zones
        green_zone = data[(data["Impact"] >= 5) & (data["Complexity"] <= 4)].shape[0]
        orange_zone = data[(data["Impact"] <= 4) & (data["Complexity"] >= 5)].shape[0]
        yellow_zone = len(data) - green_zone - orange_zone
        
        with col1:
            st.markdown(f"""
            <div style='background-color: #e3f6ff; padding: 10px; border-radius: 5px;'>
                <h3 style='text-align: center; margin: 0;'>{len(data)}</h3>
                <p style='text-align: center; margin: 0;'>Total Opportunities</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            avg_score = data['Score'].mean() if not data.empty else 0
            st.markdown(f"""
            <div style='background-color: #f5f5f5; padding: 10px; border-radius: 5px;'>
                <h3 style='text-align: center; margin: 0;'>{avg_score:.2f}</h3>
                <p style='text-align: center; margin: 0;'>Average Score</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
            <div style='background-color: #c6f5d3; padding: 10px; border-radius: 5px;'>
                <h3 style='text-align: center; margin: 0;'>{green_zone}</h3>
                <p style='text-align: center; margin: 0;'>High Priority</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col4:
            st.markdown(f"""
            <div style='background-color: #f8d3d3; padding: 10px; border-radius: 5px;'>
                <h3 style='text-align: center; margin: 0;'>{orange_zone}</h3>
                <p style='text-align: center; margin: 0;'>Low Priority</p>
            </div>
            """, unsafe_allow_html=True)
        
        # Status Distribution
        st.subheader("Status Distribution")
        status_counts = data["Status"].value_counts().reset_index()
        status_counts.columns = ["Status", "Count"]
        
        status_chart = alt.Chart(status_counts).mark_bar().encode(
            x=alt.X('Status:N', sort=STATUS_OPTIONS),
            y='Count:Q',
            color=alt.Color('Status:N', scale=alt.Scale(scheme='category10'))
        ).properties(height=200)
        
        st.altair_chart(status_chart, use_container_width=True)
        
        # Top opportunities
        st.subheader("Top Opportunities")
        
        # Tabs for different views
        tab1, tab2, tab3 = st.tabs(["By Score", "By Impact", "Recent Updates"])
        
        with tab1:
            top_scores = data.nlargest(5, 'Score')[['ID', 'Opportunity', 'Score', 'Status']]
            st.dataframe(top_scores, use_container_width=True, hide_index=True)
        
        with tab2:
            high_impact = data.nlargest(5, 'Impact')[['ID', 'Opportunity', 'Impact', 'Status']]
            st.dataframe(high_impact, use_container_width=True, hide_index=True)
        
        with tab3:
            recent = data.sort_values('Modified', ascending=False).head(5)[['ID', 'Opportunity', 'Modified', 'Status']]
            st.dataframe(recent, use_container_width=True, hide_index=True)
        
        # Progress tracking
        st.subheader("Progress Tracking")
        
        # Calculate progress percentages by status
        total = len(data)
        deployed_pct = len(data[data["Status"] == "Deployed"]) / total * 100 if total > 0 else 0
        development_pct = len(data[data["Status"] == "In development"]) / total * 100 if total > 0 else 0
        validated_pct = len(data[data["Status"] == "Validated"]) / total * 100 if total > 0 else 0
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown("**Deployed**")
            st.progress(deployed_pct / 100)
            st.text(f"{deployed_pct:.1f}%")
        
        with col2:
            st.markdown("**In Development**")
            st.progress(development_pct / 100)
            st.text(f"{development_pct:.1f}%")
        
        with col3:
            st.markdown("**Validated**")
            st.progress(validated_pct / 100)
            st.text(f"{validated_pct:.1f}%")
    
    else:
        st.info("No data available. Add opportunities in the Management page.")
        if st.button("Go to Management Page", on_click=lambda: st.session_state.update({"page": "management"})):
            st.session_state.update({"page": "management"})

def management_page(data, lang="en"):
    """Render the management page"""
    
    # Add new opportunity
    st.header("➕ " + get_text("add_new", lang))
    
    with st.form("add_form", clear_on_submit=True):
        col1, col2 = st.columns(2)
        
        with col1:
            opportunity = st.text_area("Opportunity", height=80)
            
            # Use custom areas if available
            custom_areas = st.session_state.config.get("custom_areas", [])
            if custom_areas:
                area = st.selectbox("Area", options=[""] + custom_areas)
            else:
                area = st.text_input("Area")
                
            related_to = st.text_input("Related to")
            type_ = st.selectbox("Type", TYPE_OPTIONS)
        
        with col2:
            # Use custom topics if available
            custom_topics = st.session_state.config.get("custom_topics", [])
            if custom_topics:
                topic = st.selectbox("Topic", options=[""] + custom_topics)
            else:
                topic = st.text_input("Topic")
                
            impact = st.slider("Impact", 0.0, 10.0, value=5.0, step=0.1, 
                              help="Higher values indicate higher potential impact")
            complexity = st.slider("Complexity", 0.0, 10.0, value=5.0, step=0.1,
                                  help="Higher values indicate higher implementation complexity")
            score = compute_score(impact, complexity)
            
            # Display score with a visual indicator
            col_score, col_viz = st.columns([1, 1])
            with col_score:
                st.metric("Score", score)
            with col_viz:
                st.pyplot(display_score_indicator(score, size=80))
            
            status = st.selectbox("Status", STATUS_OPTIONS)
        
        if st.form_submit_button("Add Opportunity"):
            if not opportunity:
                st.error("Opportunity description is required")
            else:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
                unique_id = generate_unique_id()
                
                new_row = pd.DataFrame([{
                    "ID": None,  # Will be assigned during refresh
                    "Opportunity": opportunity,
                    "Related to": related_to,
                    "Area": area,
                    "Type": type_,
                    "Topic": topic,
                    "Impact": impact,
                    "Complexity": complexity,
                    "Score": score,
                    "Status": status,
                    "Created": timestamp,
                    "Modified": timestamp,
                    "UUID": unique_id  # Added for tracking even if ID changes
                }])
                
                st.session_state.data = pd.concat([data, new_row], ignore_index=True)
                st.session_state.data = refresh_data(st.session_state.data)
                if save_data(st.session_state.data):
                    st.success("✅ Opportunity added successfully!")
                    # Show details of added item
                    st.json(new_row.to_dict('records')[0])
                    st.rerun()
    
    # Edit existing opportunity
    if not data.empty:
        st.header("✏️ " + get_text("edit", lang))
        
        # More flexible selection with search
        col1, col2 = st.columns([1, 2])
        with col1:
            select_method = st.radio("Selection method", ["ID", "Search"])
        
        if select_method == "ID":
            selected_id = st.selectbox("Select opportunity ID", sorted(data["ID"].unique()))
            row = data.loc[data["ID"] == selected_id].iloc[0]
        else:
            search_term = st.text_input("Search in opportunities", "")
            if search_term:
                filtered = data[data["Opportunity"].str.contains(search_term, case=False)]
                if not filtered.empty:
                    selected_idx = st.selectbox(
                        "Select from results", 
                        options=filtered.index,
                        format_func=lambda x: f"ID {filtered.loc[x, 'ID']}: {filtered.loc[x, 'Opportunity'][:50]}..."
                    )
                    row = data.loc[selected_idx]
                else:
                    st.warning("No results found")
                    st.stop()
            else:
                st.info("Enter search term")
                st.stop()
        
        with st.form("edit_form"):
            col1, col2 = st.columns(2)
            
            with col1:
                opportunity = st.text_area("Opportunity", value=row["Opportunity"], height=80)
                
                # Use custom areas if available
                custom_areas = st.session_state.config.get("custom_areas", [])
                if custom_areas and row["Area"] not in custom_areas:
                    custom_areas.append(row["Area"])
                
                if custom_areas:
                    area = st.selectbox("Area", options=custom_areas, index=custom_areas.index(row["Area"]) if row["Area"] in custom_areas else 0)
                else:
                    area = st.text_input("Area", value=row["Area"])
                    
                related_to = st.text_input("Related to", value=row["Related to"])
                type_ = st.selectbox("Type", TYPE_OPTIONS, index=TYPE_OPTIONS.index(row["Type"]) if row["Type"] in TYPE_OPTIONS else 0)
            
            with col2:
                # Use custom topics if available
                custom_topics = st.session_state.config.get("custom_topics", [])
                if custom_topics and row["Topic"] not in custom_topics:
                    custom_topics.append(row["Topic"])
                
                if custom_topics:
                    topic = st.selectbox("Topic", options=custom_topics, index=custom_topics.index(row["Topic"]) if row["Topic"] in custom_topics else 0)
                else:
                    topic = st.text_input("Topic", value=row["Topic"])
                
                impact = st.slider("Impact", 0.0, 10.0, value=float(row["Impact"]), step=0.1)
                complexity = st.slider("Complexity", 0.0, 10.0, value=float(row["Complexity"]), step=0.1)
                score = compute_score(impact, complexity)
                
                # Display score with a visual indicator and delta
                col_score, col_viz = st.columns([1, 1])
                with col_score:
                    score_delta = round(score - float(row["Score"]), 1)
                    st.metric("Score", f"{score:.1f}", 
                             delta=score_delta if score_delta != 0 else None,
                             delta_color="normal")
                
                with col_viz:
                    score_fig = display_score_indicator(score, size=100)
                    st.pyplot(score_fig)
                
                # Add score interpretation
                if score >= 8:
                    st.markdown("**High priority opportunity!** 🚀")
                elif score >= 6:
                    st.markdown("**Medium priority opportunity.** ⏳")
                else:
                    st.markdown("**Low priority opportunity.** 📋")
                
                status = st.selectbox("Status", STATUS_OPTIONS, 
                                     index=STATUS_OPTIONS.index(row["Status"]) if row["Status"] in STATUS_OPTIONS else 0)
            
            if st.form_submit_button("Update"):
                if not opportunity:
                    st.error("Opportunity description is required")
                else:
                    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
                    mask = st.session_state.data["ID"] == row["ID"]
                    
                    # Update all fields
                    st.session_state.data.loc[mask, "Opportunity"] = opportunity
                    st.session_state.data.loc[mask, "Related to"] = related_to
                    st.session_state.data.loc[mask, "Area"] = area
                    st.session_state.data.loc[mask, "Type"] = type_
                    st.session_state.data.loc[mask, "Topic"] = topic
                    st.session_state.data.loc[mask, "Impact"] = impact
                    st.session_state.data.loc[mask, "Complexity"] = complexity
                    st.session_state.data.loc[mask, "Score"] = score
                    st.session_state.data.loc[mask, "Status"] = status
                    st.session_state.data.loc[mask, "Modified"] = timestamp
                    
                    st.session_state.data = refresh_data(st.session_state.data)
                    if save_data(st.session_state.data):
                        st.success(f"✅ Opportunity ID {row['ID']} updated!")
                        st.rerun()
    
    # Delete opportunity
    if not data.empty:
        st.header("❌ " + get_text("delete", lang))
        
        delete_id = st.selectbox("Select opportunity to delete", 
                                sorted(data["ID"].unique()),
                                format_func=lambda x: f"ID {x}: {data.loc[data['ID'] == x, 'Opportunity'].values[0][:50]}...")
        
        col1, col2 = st.columns([1, 3])
        with col1:
            # Add confirmation checkbox for safety
            confirm = st.checkbox(get_text("confirm_delete", lang))
            
            if st.button("🗑️ Delete", type="primary", disabled=not confirm):
                if confirm:
                    # Create backup before deletion
                    create_backup(st.session_state.data)
                    
                    # Perform deletion
                    st.session_state.data = st.session_state.data[st.session_state.data["ID"] != delete_id]
                    st.session_state.data = refresh_data(st.session_state.data)
                    if save_data(st.session_state.data):
                        st.success(f"🗑️ Opportunity ID {delete_id} deleted.")
                        st.rerun()

def visualization_page(data, lang="en"):
    """Render the visualization page"""
    st.header("📊 Visual Analysis")
    
    # Filters in a more compact form
    with st.expander("🔍 Filters", expanded=True):
        col1, col2, col3 = st.columns(3)
        
        with col1:
            types = st.multiselect(
                "Filter by Type", 
                data["Type"].dropna().unique() if not data.empty else [],
                default=data["Type"].dropna().unique() if not data.empty else []
            )
        
        with col2:
            areas = st.multiselect(
                "Filter by Area", 
                data["Area"].dropna().unique() if not data.empty else [],
                default=data["Area"].dropna().unique() if not data.empty else []
            )
        
        with col3:
            status_filter = st.multiselect(
                "Filter by Status", 
                data["Status"].dropna().unique() if not data.empty else [],
                default=data["Status"].dropna().unique() if not data.empty else []
            )
    
    # Apply filters
    if not data.empty:
        filtered_data = data.copy()
        
        if types:
            filtered_data = filtered_data[filtered_data["Type"].isin(types)]
        if areas:
            filtered_data = filtered_data[filtered_data["Area"].isin(areas)]
        if status_filter:
            filtered_data = filtered_data[filtered_data["Status"].isin(status_filter)]
    else:
        filtered_data = data
    
    # Display filtered data with improved formatting
    st.subheader("📋 Filtered Opportunities")
    if not filtered_data.empty:
        # Color coding function
        def color_score(val):
            if val > 8:
                return 'background-color: #c6f5d3'
            elif val >= 6:
                return 'background-color: #fff3b3'
            else:
                return 'background-color: #f8d3d3'
        
        # Add emoji indicators for status
        def add_status_emoji(status):
            emoji_map = {
                "Idea": "💡",
                "To explore": "🔍",
                "Validated": "✅",
                "In development": "🛠️",
                "Deployed": "🚀"
            }
            return f"{emoji_map.get(status, '')} {status}"
        
        display_df = filtered_data.copy()
        display_df["Status"] = display_df["Status"].apply(add_status_emoji)
        
        styled_df = display_df.style.applymap(color_score, subset=["Score"])
        st.dataframe(styled_df, use_container_width=True, hide_index=True)
    else:
        st.info("No data matching the filters")
    
    # Interactive Heatmap with tabs for different visualizations
    if not filtered_data.empty:
        chart_tabs = st.tabs(["Impact vs Complexity Heatmap", "Priority Distribution", "Status Breakdown"])
        
        with chart_tabs[0]:
            st.subheader("📊 Impact vs Complexity Heatmap")
            
            # Improved heatmap
            fig, ax = plt.subplots(figsize=(10, 8))
            ax.set_xlim(-0.5, 10.5)
            ax.set_ylim(-0.5, 10.5)
            ax.set_xlabel("Complexity", fontsize=12)
            ax.set_ylabel("Impact", fontsize=12)
            ax.set_title("Impact vs Complexity Heatmap", fontsize=14, fontweight='bold')
            ax.set_xticks(range(11))
            ax.set_yticks(range(11))
            ax.grid(True, linestyle='--', alpha=0.3)
            
            # Background zones with better coloring and labels
            # Green zone - high impact, low complexity (prioritize)
            rect1 = patches.Rectangle((0, 5), 5, 5, facecolor='#A8D5BA', alpha=0.3, edgecolor='none')
            ax.add_patch(rect1)
            ax.text(2.5, 7.5, "HIGH PRIORITY", ha='center', va='center', fontweight='bold', alpha=0.5, fontsize=14)
            
            # Orange zone - low impact, high complexity (deprioritize)
            rect2 = patches.Rectangle((5, 0), 5, 5, facecolor='#F9C6A8', alpha=0.3, edgecolor='none')
            ax.add_patch(rect2)
            ax.text(7.5, 2.5, "LOW PRIORITY", ha='center', va='center', fontweight='bold', alpha=0.5, fontsize=14)
            
            # Yellow zone - everything else (evaluate)
            rect3 = patches.Rectangle((0, 0), 5, 5, facecolor='#FFF5B1', alpha=0.3, edgecolor='none')
            ax.add_patch(rect3)
            rect4 = patches.Rectangle((5, 5), 5, 5, facecolor='#FFF5B1', alpha=0.3, edgecolor='none')
            ax.add_patch(rect4)
            
            # Plot opportunities with jitter to avoid overlap
            jitter_x = filtered_data["Complexity"] + np.random.uniform(-0.15, 0.15, len(filtered_data))
            jitter_y = filtered_data["Impact"] + np.random.uniform(-0.15, 0.15, len(filtered_data))
            
            # Use color coding by status
            status_colors = {
                "Idea": "lightgray",
                "To explore": "skyblue",
                "Validated": "orange",
                "In development": "purple",
                "Deployed": "green"
            }
            
            # Default color for any status not in the dictionary
            colors = [status_colors.get(status, "gray") for status in filtered_data["Status"]]
            
            # Size based on score
            #sizes = filtered_data["Score"] * 200
            sizes = 1000

            scatter = ax.scatter(
                jitter_x, jitter_y, 
                s=sizes, 
                c=colors,
                edgecolors='black', 
                linewidth=1.2, 
                alpha=0.9
            )
            
            # Add ID labels
            for i, (x, y, id_val) in enumerate(zip(jitter_x, jitter_y, filtered_data["ID"])):
                ax.text(x, y, str(int(id_val)), 
                       fontsize=9, ha='center', va='center', 
                       weight='bold', color='white')
            
            # Add legend for status colors
            from matplotlib.lines import Line2D
            legend_elements = [
                Line2D([0], [0], marker='o', color='w', markerfacecolor=color, 
                       label=status, markersize=10)
                for status, color in status_colors.items() 
                if status in filtered_data["Status"].values
            ]
            
            ax.legend(handles=legend_elements, title="Status", 
                     loc='upper left', bbox_to_anchor=(1, 1))
            
            plt.tight_layout()
            st.pyplot(fig)
            
            # Download heatmap
            buf = io.BytesIO()
            fig.savefig(buf, format='png', dpi=300, bbox_inches='tight')
            buf.seek(0)
            
            st.download_button(
                "🖼️ Download Heatmap",
                data=buf.getvalue(),
                file_name=f"heatmap_{datetime.now().strftime('%Y%m%d_%H%M')}.png",
                mime="image/png"
            )
        
        with chart_tabs[1]:
            st.subheader("Priority Distribution")
            
            # Create a priority category
            filtered_data['Priority'] = pd.cut(
                filtered_data['Score'],
                bins=[0, 4, 6, 8, 10],
                labels=['Low', 'Medium-Low', 'Medium-High', 'High']
            )
            
            # Create bar chart of priorities
            priority_counts = filtered_data['Priority'].value_counts().reset_index()
            priority_counts.columns = ['Priority', 'Count']
            
            # Order for the priorities
            order = ['High', 'Medium-High', 'Medium-Low', 'Low']
            priority_counts['Priority'] = pd.Categorical(
                priority_counts['Priority'], 
                categories=order, 
                ordered=True
            )
            priority_counts = priority_counts.sort_values('Priority')
            
            # Create bar chart
            priority_chart = alt.Chart(priority_counts).mark_bar().encode(
                x=alt.X('Priority:N', sort=order),
                y='Count:Q',
                color=alt.Color('Priority:N', scale=alt.Scale(
                    domain=['High', 'Medium-High', 'Medium-Low', 'Low'],
                    range=['#c6f5d3', '#d4f5b3', '#fff3b3', '#f8d3d3']
                ))
            ).properties(height=300)
            
            st.altair_chart(priority_chart, use_container_width=True)
        
        with chart_tabs[2]:
            st.subheader("Status Breakdown")
            
            # Create donut chart for status
            status_counts = filtered_data['Status'].value_counts().reset_index()
            status_counts.columns = ['Status', 'Count']
            
            fig, ax = plt.subplots(figsize=(8, 8))
            
            # Custom colors for each status
            status_colors_pie = {
                "Idea": "#aaaaaa",
                "To explore": "#6baed6",
                "Validated": "#fd8d3c",
                "In development": "#9e9ac8",
                "Deployed": "#74c476"
            }
            
            # Map colors to the statuses in our data
            colors = [status_colors_pie.get(s, "#999999") for s in status_counts['Status']]
            
            # Create donut chart
            wedges, texts, autotexts = ax.pie(
                status_counts['Count'], 
                labels=status_counts['Status'],
                colors=colors,
                autopct='%1.1f%%',
                startangle=90,
                wedgeprops={'edgecolor': 'white', 'linewidth': 2}
            )
            
            # Make donut hole
            circle = plt.Circle((0, 0), 0.4, fc='white')
            ax.add_artist(circle)
            
            # Equal aspect ratio ensures that pie is drawn as a circle
            ax.set_aspect('equal')
            
            # Style the text
            for text in texts:
                text.set_fontsize(12)
                text.set_fontweight('bold')
            
            for autotext in autotexts:
                autotext.set_fontsize(10)
                autotext.set_fontweight('bold')
                autotext.set_color('white')
            
            ax.set_title('Status Distribution', fontsize=16, pad=20)
            
            st.pyplot(fig)
            
            # Add a table with status counts and percentages
            status_table = status_counts.copy()
            status_table['Percentage'] = (status_table['Count'] / status_table['Count'].sum() * 100).round(1)
            status_table['Percentage'] = status_table['Percentage'].astype(str) + '%'
            
            st.dataframe(status_table, use_container_width=True, hide_index=True)

def export_page(data, lang="en"):
    """Render the export page"""
    st.header("⬇️ Export Data")
    
    # Apply filters for export
    if not data.empty:
        col1, col2 = st.columns(2)
        
        with col1:
            types = st.multiselect(
                "Types to export", 
                data["Type"].dropna().unique(),
                default=data["Type"].dropna().unique()
            )
        
        with col2:
            status_filter = st.multiselect(
                "Status to export", 
                data["Status"].dropna().unique(),
                default=data["Status"].dropna().unique()
            )
        
        # More granular area selection
        all_areas = data["Area"].dropna().unique()
        select_all_areas = st.checkbox("Select all areas", value=True)
        
        if select_all_areas:
            areas = all_areas
        else:
            areas = st.multiselect(
                "Areas to export", 
                all_areas,
                default=[]
            )
        
        export_data = data[
            data["Type"].isin(types) & 
            data["Area"].isin(areas) & 
            data["Status"].isin(status_filter)
        ]
    else:
        export_data = data
    
    if not export_data.empty:
        st.write(f"Exporting {len(export_data)} opportunities")
        
        # Export options
        export_tabs = st.tabs(["CSV", "Excel", "JSON", "Preview"])
        
        with export_tabs[0]:
            csv_data = export_data.to_csv(index=False).encode('utf-8')
            
            col1, col2 = st.columns([1, 2])
            with col1:
                st.download_button(
                    "📄 Download CSV",
                    data=csv_data,
                    file_name=f"opportunities_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                    mime="text/csv"
                )
            
            with col2:
                st.info("CSV format is compatible with most spreadsheet applications.")
        
        with export_tabs[1]:
            # Excel export with better formatting
            output = io.BytesIO()
            
            try:
                import openpyxl
                from openpyxl.styles import PatternFill, Font, Alignment
                
                # Create a writer with openpyxl
                writer = pd.ExcelWriter(output, engine='openpyxl')
                export_data.to_excel(writer, index=False, sheet_name='Opportunities')
                
                # Get the openpyxl workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Opportunities']
                
                # Define fills for different score ranges
                high_fill = PatternFill(start_color="C6F5D3", end_color="C6F5D3", fill_type="solid")
                med_fill = PatternFill(start_color="FFF3B3", end_color="FFF3B3", fill_type="solid")
                low_fill = PatternFill(start_color="F8D3D3", end_color="F8D3D3", fill_type="solid")
                
                # Format header row
                header_font = Font(bold=True, size=12)
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                # Color code based on score
                score_col = COLUMNS.index("Score") + 1  # +1 because openpyxl is 1-indexed
                
                for row in range(2, len(export_data) + 2):  # +2 to account for header and 1-indexing
                    score_cell = worksheet.cell(row=row, column=score_col)
                    score_value = score_cell.value
                    
                    if score_value > 8:
                        score_cell.fill = high_fill
                    elif score_value >= 6:
                        score_cell.fill = med_fill
                    else:
                        score_cell.fill = low_fill
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    # Limit width to 50 characters
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Save the workbook
                writer.close()
                
                col1, col2 = st.columns([1, 2])
                with col1:
                    st.download_button(
                        "📊 Download Excel",
                        data=output.getvalue(),
                        file_name=f"opportunities_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                with col2:
                    st.info("Excel format includes formatting and color-coding.")
            
            except ImportError:
                # Simplified Excel export if openpyxl is not available
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    export_data.to_excel(writer, index=False, sheet_name='Opportunities')
                
                st.download_button(
                    "📊 Download Excel",
                    data=output.getvalue(),
                    file_name=f"opportunities_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        with export_tabs[2]:
            # JSON export
            json_data = export_data.to_json(orient='records', indent=2)
            
            col1, col2 = st.columns([1, 2])
            with col1:
                st.download_button(
                    "📋 Download JSON",
                    data=json_data,
                    file_name=f"opportunities_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
                    mime="application/json"
                )
            
            with col2:
                st.info("JSON format is useful for API integrations and data processing.")
        
        with export_tabs[3]:
            # Preview the data being exported
            st.subheader("Export Preview")
            st.dataframe(export_data, use_container_width=True)
            
            # Summary stats
            st.subheader("Summary Statistics")
            summary = pd.DataFrame({
                'Total Opportunities': [len(export_data)],
                'Average Score': [export_data['Score'].mean()],
                'Highest Score': [export_data['Score'].max()],
                'Lowest Score': [export_data['Score'].min()],
                'High Priority Count': [len(export_data[export_data['Score'] >= 8])],
                'Medium Priority Count': [len(export_data[(export_data['Score'] >= 6) & (export_data['Score'] < 8)])],
                'Low Priority Count': [len(export_data[export_data['Score'] < 6])]
            })
            
            st.dataframe(summary.T, use_container_width=True)
    else:
        st.info("No data to export")

def settings_page(lang="en"):
    """Render the settings page"""
    st.header("⚙️ Settings")
    
    config = st.session_state.config
    
    with st.form("settings_form"):
        # Language selection
        lang_options = {
            "en": "English",
            "fr": "Français"
        }
        
        selected_lang = st.selectbox(
            "Language / Langue",
            options=list(lang_options.keys()),
            format_func=lambda x: lang_options[x],
            index=list(lang_options.keys()).index(config.get("language", "en"))
        )
        
        # Theme selection
        theme_options = {
            "light": "Light",
            "dark": "Dark"
        }
        
        selected_theme = st.selectbox(
            "Theme",
            options=list(theme_options.keys()),
            format_func=lambda x: theme_options[x],
            index=list(theme_options.keys()).index(config.get("theme", "light"))
        )
        
        # Backup settings
        backup_frequency = st.slider(
            "Backup frequency (number of changes before auto-backup)",
            min_value=1,
            max_value=20,
            value=config.get("backup_frequency", 5)
        )
        
        # Custom area management
        st.subheader("Custom Areas")
        custom_areas_str = st.text_area(
            "Custom Areas (one per line)",
            value="\n".join(config.get("custom_areas", [])),
            height=100,
            help="Enter custom areas, one per line. These will be available as dropdown options."
        )
        
        # Custom topic management
        st.subheader("Custom Topics")
        custom_topics_str = st.text_area(
            "Custom Topics (one per line)",
            value="\n".join(config.get("custom_topics", [])),
            height=100,
            help="Enter custom topics, one per line. These will be available as dropdown options."
        )
        
        if st.form_submit_button("Save Settings"):
            # Process custom areas and topics
            custom_areas = [area.strip() for area in custom_areas_str.split("\n") if area.strip()]
            custom_topics = [topic.strip() for topic in custom_topics_str.split("\n") if topic.strip()]
            
            # Update config
            config.update({
                "language": selected_lang,
                "theme": selected_theme,
                "backup_frequency": backup_frequency,
                "custom_areas": custom_areas,
                "custom_topics": custom_topics
            })
            
            # Save to session state and file
            st.session_state.config = config
            save_config(config)
            st.success("✅ Settings saved successfully!")
            st.rerun()
    
    # Backup management
    st.header("🔄 Backup Management")
    
    # List available backups
    backups = sorted(Path(BACKUP_FOLDER).glob("*.csv"), key=os.path.getmtime, reverse=True)
    
    if backups:
        backup_options = {}
        for backup in backups:
            timestamp = datetime.fromtimestamp(os.path.getmtime(backup)).strftime("%Y-%m-%d %H:%M:%S")
            backup_options[str(backup)] = f"{timestamp} ({backup.name})"
        
        selected_backup = st.selectbox(
            "Available backups",
            options=list(backup_options.keys()),
            format_func=lambda x: backup_options[x]
        )
        
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("Restore Selected Backup"):
                # Load the backup
                backup_df = pd.read_csv(selected_backup)
                
                # Create a backup of current data first
                if Path(DATA_FILE).exists():
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    current_backup = os.path.join(BACKUP_FOLDER, f"pre_restore_backup_{timestamp}.csv")
                    st.session_state.data.to_csv(current_backup, index=False)
                
                # Restore the backup
                st.session_state.data = backup_df
                save_data(st.session_state.data)
                st.success("✅ Backup restored successfully!")
                st.rerun()
        
        with col2:
            if st.button("Download Selected Backup"):
                with open(selected_backup, "rb") as f:
                    backup_data = f.read()
                
                st.download_button(
                    "📥 Download Backup",
                    data=backup_data,
                    file_name=os.path.basename(selected_backup),
                    mime="text/csv"
                )
    else:
        st.info("No backups available")
    
    # Restore from file upload
    st.subheader("Restore from File")
    
    uploaded_file = st.file_uploader("Upload a backup CSV file", type=["csv"])
    
    if uploaded_file is not None:
        try:
            # Create a backup of current data first
            if Path(DATA_FILE).exists() and not st.session_state.data.empty:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                current_backup = os.path.join(BACKUP_FOLDER, f"pre_upload_backup_{timestamp}.csv")
                st.session_state.data.to_csv(current_backup, index=False)
            
            # Load the uploaded file
            uploaded_df = pd.read_csv(uploaded_file)
            
            # Validate columns
            missing_cols = [col for col in COLUMNS if col not in uploaded_df.columns]
            
            if missing_cols:
                st.error(f"The uploaded file is missing required columns: {', '.join(missing_cols)}")
            else:
                # Preview the data
                st.subheader("Preview of uploaded data")
                st.dataframe(uploaded_df.head(), use_container_width=True)
                
                if st.button("Confirm Restore from Uploaded File"):
                    st.session_state.data = uploaded_df
                    save_data(st.session_state.data)
                    st.success("✅ Data restored from uploaded file!")
                    st.rerun()
        
        except Exception as e:
            st.error(f"Error processing uploaded file: {e}")
    
    # Advanced settings
    with st.expander("⚠️ Advanced Settings"):
        st.warning("These actions cannot be undone. Use with caution.")
        
        st.markdown("### 🔄 Reset All Data")
        st.markdown("This will delete all opportunities from all dashboards and reset the application to its initial state.")
        
        reset_col1, reset_col2 = st.columns([1, 2])
        
        with reset_col1:
            reset_confirmation = st.checkbox("I understand this will permanently delete ALL data")
        
        with reset_col2:
            if st.button("🔄 Reset All Data", type="primary", disabled=not reset_confirmation):
                if reset_confirmation:
                    # Create one final backup before reset
                    if Path(DATA_FILE).exists() and not st.session_state.data.empty:
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        final_backup = os.path.join(BACKUP_FOLDER, f"final_backup_before_reset_{timestamp}.csv")
                        st.session_state.data.to_csv(final_backup, index=False)
                    
                    # Reset the data
                    st.session_state.data = pd.DataFrame(columns=COLUMNS)
                    if Path(DATA_FILE).exists():
                        Path(DATA_FILE).unlink()
                    
                    # Reset other session state variables
                    st.session_state.save_counter = 0
                    
                    st.success("🔄 All data has been completely reset. A backup was created before deletion.")
                    
                    # Create a download link for the backup
                    if Path(final_backup).exists():
                        with open(final_backup, "rb") as f:
                            backup_data = f.read()
                        
                        st.download_button(
                            "📥 Download Backup of Deleted Data",
                            data=backup_data,
                            file_name=os.path.basename(final_backup),
                            mime="text/csv"
                        )
                    
                    # Add a short delay before rerun to ensure the user sees the success message
                    import time
                    time.sleep(1)
                    st.rerun()

# ----------- Main App Flow -----------
def main():
    """Main application flow"""
    # Initialize folders
    initialize_folders()
    
    # Load config if not in session state
    if "config" not in st.session_state:
        st.session_state.config = load_config()
    
    # Set theme based on config
    theme = st.session_state.config.get("theme", "light")
    
    # Access language from config
    lang = st.session_state.config.get("language", "en")
    
    # Set page config
    st.set_page_config(
        page_title=get_text("app_title", lang),
        layout="wide",
        initial_sidebar_state="expanded",
        menu_items={
            'About': f"AI Opportunities Priority Heat Map v{APP_VERSION}"
        }
    )
    
    # Initialize data before any operations
    if "data" not in st.session_state:
        st.session_state.data = load_data()
    
    if "save_counter" not in st.session_state:
        st.session_state.save_counter = 0
    
    # Initialize or validate page value
    if "page" not in st.session_state:
        st.session_state.page = "dashboard"
    
    # Validate that the current page exists in our valid pages list
    valid_pages = ["dashboard", "management", "visualization", "export", "settings"]
    if st.session_state.page not in valid_pages:
        st.session_state.page = "dashboard"  # Reset to dashboard if invalid page
    
    # Add custom CSS
    st.markdown("""
        <style>
            /* General styling */
            .stApp {
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            }
            
            /* Card styling */
            div[data-testid="stMetricValue"] {
                font-size: 2rem !important;
                font-weight: bold;
            }
            
            /* Improve sidebar styling */
            .css-1d391kg {
                padding-top: 2rem;
            }
            
            /* Button styling */
            .stButton button {
                border-radius: 6px;
            }
            
            /* Title styling */
            h1 {
                margin-bottom: 1.5rem !important;
            }
            
            /* Improve metric cards */
            [data-testid="stMetricValue"] > div {
                background-color: #f0f2f6;
                padding: 0.5rem;
                border-radius: 5px;
            }
        </style>
    """, unsafe_allow_html=True)
    
    # Title
    st.markdown(f"""
        <h1 style='text-align: center; width: 100%; font-size: 2.5rem;'>
            {get_text("app_title", lang)}
        </h1>
    """, unsafe_allow_html=True)
    
    # Add version info and last update timestamp
    col1, col2 = st.columns([3, 1])
    with col2:
        st.markdown(f"<div style='text-align: right; font-size: 0.8rem; color: gray;'>v{APP_VERSION}</div>", 
                  unsafe_allow_html=True)
    
    # Sidebar
    with st.sidebar:
        # Add logo (if exists)
        add_logo()
        
        st.header(get_text("navigation", lang))
        
        # Define page mapping - using consistent keys between pages dictionary and session state
        # Use lowercase consistently for keys
        pages = {
            "dashboard": get_text("dashboard", lang),
            "management": get_text("management", lang),  # Changed from "Gestion" to "management"
            "visualization": get_text("visualization", lang),
            "export": get_text("export", lang),
            "settings": get_text("settings", lang)
        }
        
        selected_page = st.selectbox(
            "Page",
            options=list(pages.keys()),
            format_func=lambda x: pages[x],
            index=list(pages.keys()).index(st.session_state.page)
        )
        
        # Update session state with selected page
        st.session_state.page = selected_page
        
        # Display some stats in sidebar
        if not st.session_state.data.empty:
            st.divider()
            st.subheader("Quick Stats")
            st.markdown(f"**Total opportunities:** {len(st.session_state.data)}")
            
            # Opportunity count by status
            status_counts = st.session_state.data["Status"].value_counts()
            for status in STATUS_OPTIONS:
                if status in status_counts:
                    st.markdown(f"**{status}:** {status_counts[status]}")
            
            # Last modification
            if "Modified" in st.session_state.data.columns:
                last_mod = st.session_state.data["Modified"].max() if not st.session_state.data.empty else "Never"
                st.markdown(f"**Last update:** {last_mod}")
    
    # Render the selected page
    if selected_page == "dashboard":
        dashboard_page(st.session_state.data, lang)
    elif selected_page == "management":
        management_page(st.session_state.data, lang)
    elif selected_page == "visualization":
        visualization_page(st.session_state.data, lang)
    elif selected_page == "export":
        export_page(st.session_state.data, lang)
    elif selected_page == "settings":
        settings_page(lang)

if __name__ == "__main__":
    main()